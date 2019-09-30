import openpyxl

def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getColCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readData(file,sheetName,rowNum,columnNo):
    workbook=openpyxl.load_workbook(sheetName)
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNum,col=columnNo).value

def writeData(file,sheetName,rowNum,columnNo,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNum,col=columnNo).value=data
    workbook.save(file)


