import openpyxl

def getRowCount(file,sheetName):
    workBook=openpyxl.load_workbook(file)
    sheet=workBook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColCount(file,sheetName):
    workBook=openpyxl.load_workbook(file)
    sheet=workBook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnNo):
    workBook=openpyxl.load_workbook(file)
    sheet=workBook.get_sheet_by_name(sheetName)
    return  sheet.cell(row=rownum,column=columnNo).value

def writeData(file,sheetName,rownum,columnNo,data):
    workBook=openpyxl.load_workbook(file)
    sheet=workBook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum,column=columnNo).value=data
    workBook.save(file)